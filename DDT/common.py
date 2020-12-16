import openpyxl

class Common():

    def __init__(self, FilePath, SheetName):
        self.workbook = openpyxl.load_workbook(FilePath)  # Load xlsx into workbook variable
        self.sheet = self.workbook[SheetName]  # define which sheet to use

    def fetch_row_count(self):
        rows = self.sheet.max_row  # read how many rows are in the file
        return rows

    def fetch_column_count(self):
        columns = self.sheet.max_column
        return columns

    def fetch_key_names(self):
        c = self.sheet.max_column
        li = []
        for i in range(1, c+1):
            cell = self.sheet.cell(row=1, column=i)
            li.append(cell.value)

        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = self.sheet.max_column
        for i in range(1, c+1):
            cell = self.sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest