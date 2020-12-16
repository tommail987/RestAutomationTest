import requests
import json
from jsonpath import jsonpath
import openpyxl

def test_add_new_student():
    #API
    url = 'http://thetestingworldapi.com/api/studentsDetails'
    file_path = 'C:\\Users\\TT\\Desktop\\newStudent.json'
    f = open(file_path, 'r')
    # XLSX
    workbook = openpyxl.load_workbook('C:\\Users\\TT\\Desktop\\Students.xlsx') #Load xlsx into workbook variable
    sheet = workbook['Arkusz1'] #define which sheet to use
    rows = sheet.max_row #read how many rows are in the file
    json_file = json.loads(f.read()) #load json file into variable in json (dictionary) format
    for i in range(2, rows+1): #for each row in range, without first row because there are headers
        cell_first_name = sheet.cell(row=i, column=1) # points out which is first name cell
        cell_mid_name = sheet.cell(row=i, column=2) # points out which is middle name cell
        cell_last_name = sheet.cell(row=i, column=3) # points out which is last name cell
        cell_dob = sheet.cell(row=i, column=4) # points out which is date of birth cell

        json_file['first_name'] = cell_first_name.value # insert value into json dictionary file
        json_file['middle_name'] = cell_mid_name.value # insert value into json dictionary file
        json_file['last_name'] = cell_last_name.value # insert value into json dictionary file
        json_file['date_of_birth'] = cell_dob.value # insert value into json dictionary file

        response = requests.post(url, json_file)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201