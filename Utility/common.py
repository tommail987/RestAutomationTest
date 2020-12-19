import openpyxl
import json

class Common():

    def open_workbook(self, FilePath, SheetName):
        workbook = openpyxl.load_workbook(FilePath)  # Load xlsx into workbook variable
        sheet = self.workbook[SheetName]  # define which sheet to use
        return workbook, sheet


    def fetch_row_count(self):
        rows = self.sheet.max_row  # read how many rows are in the file
        return rows

    def fetch_column_count(self):
        columns = self.sheet.max_column
        return columns

    def fetch_key_names(self):
        c = self.sheet.max_column
        li = []
        for i in range(1, c+1):
            cell = self.sheet.cell(row=1, column=i)
            li.append(cell.value)

        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = self.sheet.max_column
        for i in range(1, c+1):
            cell = self.sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest

    def request_file(self, request_file_path):
        request_file = open(request_file_path, 'r')
        json_file = json.loads(request_file.read())
        return json_file